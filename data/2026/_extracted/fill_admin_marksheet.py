import csv
from pathlib import Path

from openpyxl import load_workbook


ADMIN = Path("data/2026/marksheet25-26_CCGL9065_EW.xlsx")
ROSTER = Path("data/2026/grading_2026_roster_base.xlsx")
OUTPUT = Path("data/2026/marksheet25-26_CCGL9065_EW_filled.xlsx")
COMPARISON = Path("data/2026/marksheet25-26_CCGL9065_EW_comparison.csv")


def headers(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def sid(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip().replace(".0", "")


def number(value):
    return round(float(value or 0), 2)


def main():
    roster_wb = load_workbook(ROSTER, data_only=True)
    grades = roster_wb["Grades"]
    rollup = roster_wb["Final_Rollup"]
    gh = headers(grades)
    rh = headers(rollup)

    rollup_by_id = {}
    for row in range(2, rollup.max_row + 1):
        student_id = sid(rollup.cell(row, rh["Student #"]).value)
        first = str(rollup.cell(row, rh["First Name"]).value or "").strip()
        last = str(rollup.cell(row, rh["Last Name"]).value or "").strip()
        rollup_by_id[student_id] = {
            "name_first_last": f"{first} {last}".strip(),
            "name_last_first": f"{last} {first}".strip(),
            "letter": rollup.cell(row, rh["Calibrated_Letter"]).value,
            "final": number(rollup.cell(row, rh["Computed_Total"]).value),
            "rank": rollup.cell(row, rh["Rank"]).value,
        }

    marks_by_id = {}
    for row in range(2, grades.max_row + 1):
        student_id = sid(grades.cell(row, gh["Student #"]).value)
        tutorial = number(grades.cell(row, gh["Tutorial"]).value)
        in_class = number(grades.cell(row, gh["Participation"]).value)
        weekly = number(grades.cell(row, gh["Weekly"]).value)
        essay = number(grades.cell(row, gh["Essay"]).value)
        collage = number(grades.cell(row, gh["Collage"]).value)
        video = number(grades.cell(row, gh["Video"]).value)

        participation = tutorial + in_class
        portfolio = essay + collage + video
        marks_by_id[student_id] = {
            "coursework": round((participation + weekly) / 40 * 100, 2),
            "exam": round(portfolio / 60 * 100, 2),
            "final": round(participation + weekly + portfolio, 2),
            "participation": participation,
            "weekly": weekly,
            "portfolio": portfolio,
        }

    admin_wb = load_workbook(ADMIN)
    sheet = admin_wb["Sheet1"]

    # Header fields that the admin template left blank.
    sheet["C6"] = "Hongshan Guo"
    sheet["C8"] = "2A"
    sheet["C9"] = "Round"

    comparison_rows = []
    seen_ids = set()
    for row in range(12, sheet.max_row + 1):
        student_id = sid(sheet.cell(row, 2).value)
        if not student_id:
            continue
        seen_ids.add(student_id)

        roll = rollup_by_id.get(student_id)
        marks = marks_by_id.get(student_id)
        if not roll or not marks:
            comparison_rows.append(
                {
                    "Row": row,
                    "UNo": student_id,
                    "Admin Name": sheet.cell(row, 3).value,
                    "Roster Name": "",
                    "Admin Grade": sheet.cell(row, 4).value,
                    "Roster Grade": "",
                    "Grade Match": "NO",
                    "Coursework-Mark": "",
                    "Exam-Mark": "",
                    "Final-Mark": "",
                    "Status": "Missing from roster workbook",
                }
            )
            continue

        sheet.cell(row, 5).value = marks["coursework"]
        sheet.cell(row, 6).value = marks["exam"]
        sheet.cell(row, 7).value = roll["final"]

        admin_grade = str(sheet.cell(row, 4).value or "").strip()
        roster_grade = str(roll["letter"] or "").strip()
        grade_match = "YES" if admin_grade == roster_grade else "NO"
        comparison_rows.append(
            {
                "Row": row,
                "UNo": student_id,
                "Admin Name": sheet.cell(row, 3).value,
                "Roster Name": roll["name_last_first"],
                "Admin Grade": admin_grade,
                "Roster Grade": roster_grade,
                "Grade Match": grade_match,
                "Coursework-Mark": marks["coursework"],
                "Exam-Mark": marks["exam"],
                "Final-Mark": roll["final"],
                "Status": "OK" if grade_match == "YES" else "Grade mismatch",
            }
        )

    for missing_id in sorted(set(rollup_by_id) - seen_ids):
        roll = rollup_by_id[missing_id]
        comparison_rows.append(
            {
                "Row": "",
                "UNo": missing_id,
                "Admin Name": "",
                "Roster Name": roll["name_last_first"],
                "Admin Grade": "",
                "Roster Grade": roll["letter"],
                "Grade Match": "NO",
                "Coursework-Mark": "",
                "Exam-Mark": "",
                "Final-Mark": roll["final"],
                "Status": "Missing from admin workbook",
            }
        )

    admin_wb.save(OUTPUT)

    fieldnames = [
        "Row",
        "UNo",
        "Admin Name",
        "Roster Name",
        "Admin Grade",
        "Roster Grade",
        "Grade Match",
        "Coursework-Mark",
        "Exam-Mark",
        "Final-Mark",
        "Status",
    ]
    with COMPARISON.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(comparison_rows)

    mismatch_count = sum(row["Status"] != "OK" for row in comparison_rows)
    print(f"wrote={OUTPUT}")
    print(f"wrote={COMPARISON}")
    print(f"checked_rows={len([r for r in comparison_rows if r['Row']])}")
    print(f"mismatches={mismatch_count}")


if __name__ == "__main__":
    main()
