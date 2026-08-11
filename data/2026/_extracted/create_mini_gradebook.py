import csv
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path("data/2026/grading_2026_roster_base.xlsx")
OUT_XLSX = Path("data/2026/grading_2026_mini_roster.xlsx")
OUT_CSV = Path("data/2026/grading_2026_mini_roster.csv")


def headers(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def sid(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    return round(float(value or 0), 2)


def main():
    source_wb = load_workbook(SOURCE, data_only=True)
    grades = source_wb["Grades"]
    rollup = source_wb["Final_Rollup"]
    gh = headers(grades)
    rh = headers(rollup)

    rollup_by_id = {
        sid(rollup.cell(row, rh["Student #"]).value): {
            "Rank": rollup.cell(row, rh["Rank"]).value,
            "Calibrated_Letter": rollup.cell(row, rh["Calibrated_Letter"]).value,
            "Computed_Total": rollup.cell(row, rh["Computed_Total"]).value,
            "Notes": rollup.cell(row, rh["Notes"]).value,
        }
        for row in range(2, rollup.max_row + 1)
    }

    rows = []
    for row in range(2, grades.max_row + 1):
        student_id = sid(grades.cell(row, gh["Student #"]).value)
        last = str(grades.cell(row, gh["Last Name"]).value or "").strip()
        first = str(grades.cell(row, gh["First Name"]).value or "").strip()
        full_name = f"{first} {last}".strip()
        tutorial = number(grades.cell(row, gh["Tutorial"]).value)
        in_class = number(grades.cell(row, gh["Participation"]).value)
        participation = round(tutorial + in_class, 2)
        assignment = number(grades.cell(row, gh["Weekly"]).value)
        essay = number(grades.cell(row, gh["Essay"]).value)
        collage = number(grades.cell(row, gh["Collage"]).value)
        video = number(grades.cell(row, gh["Video"]).value)
        portfolio = round(essay + collage + video, 2)
        total = round(participation + assignment + portfolio, 2)
        roll = rollup_by_id.get(student_id, {})

        rows.append(
            {
                "Last Name": last,
                "First Name": first,
                "Full Name": full_name,
                "Student #": student_id,
                "Email": grades.cell(row, gh["Email"]).value,
                "Participation (20%)": participation,
                "Assignment (20%)": assignment,
                "Final Exam (Portfolio) (60%)": portfolio,
                "Final Mark (100%)": number(roll.get("Computed_Total", total)),
                "Final Letter Grade": roll.get("Calibrated_Letter"),
                "Rank": roll.get("Rank"),
                "Tutorial (10%)": tutorial,
                "In-Class Participation (10%)": in_class,
                "Essay (20%)": essay,
                "Collage (20%)": collage,
                "Video (20%)": video,
                "Remarks": roll.get("Notes"),
            }
        )

    headers_out = [
        "Last Name",
        "First Name",
        "Full Name",
        "Student #",
        "Email",
        "Participation (20%)",
        "Assignment (20%)",
        "Final Exam (Portfolio) (60%)",
        "Final Mark (100%)",
        "Final Letter Grade",
        "Rank",
        "Tutorial (10%)",
        "In-Class Participation (10%)",
        "Essay (20%)",
        "Collage (20%)",
        "Video (20%)",
        "Remarks",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Mini_Grades"
    ws.append(headers_out)
    for item in rows:
        ws.append([item.get(header) for header in headers_out])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in ws.iter_rows(min_row=2, min_col=6, max_col=16):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
    for row in ws.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "0"

    mapping = wb.create_sheet("Mapping")
    mapping_rows = [
        ["Mini Column", "Source / Calculation"],
        ["Participation (20%)", "Tutorial (10%) + In-Class Participation (10%)"],
        ["Assignment (20%)", "Weekly Research & Reflection, using median of submitted weekly entries"],
        ["Final Exam (Portfolio) (60%)", "Essay (20%) + Collage (20%) + Video (20%)"],
        ["Final Mark (100%)", "Participation + Assignment + Final Exam (Portfolio)"],
        ["Final Letter Grade", "Calibrated letter from Final_Rollup in roster-base workbook"],
        ["Rank", "Rank from Final_Rollup after current calibration"],
    ]
    for row in mapping_rows:
        mapping.append(row)
    for cell in mapping[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    mapping.freeze_panes = "A2"

    for sheet in [ws, mapping]:
        for col in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(row, col).value or "") for row in range(1, sheet.max_row + 1)]
            width = min(max(12, max(len(value) for value in values) + 2), 42)
            sheet.column_dimensions[get_column_letter(col)].width = width

    wb.save(OUT_XLSX)

    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=headers_out)
        writer.writeheader()
        writer.writerows(rows)

    print(f"wrote={OUT_XLSX}")
    print(f"wrote={OUT_CSV}")
    print(f"rows={len(rows)}")


if __name__ == "__main__":
    main()
