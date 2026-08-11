from copy import copy
from pathlib import Path

from openpyxl import load_workbook


TEMPLATE = Path("marksheet24-25_CCGL9065_HGuo.xlsx")
SOURCE = Path("data/2026/grading_2026_roster_base.xlsx")
OUTPUT = Path("data/2026/marksheet25-26_CCGL9065_HGuo_filled.xlsx")


def headers(ws):
    return {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}


def sid(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    return round(float(value or 0), 2)


def copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)


def main():
    source_wb = load_workbook(SOURCE, data_only=True)
    grades = source_wb["Grades"]
    rollup = source_wb["Final_Rollup"]
    gh = headers(grades)
    rh = headers(rollup)

    rollup_by_id = {
        sid(rollup.cell(row, rh["Student #"]).value): {
            "letter": rollup.cell(row, rh["Calibrated_Letter"]).value,
            "total": number(rollup.cell(row, rh["Computed_Total"]).value),
            "rank": rollup.cell(row, rh["Rank"]).value,
        }
        for row in range(2, rollup.max_row + 1)
    }

    wb = load_workbook(TEMPLATE)
    ws = wb["Sheet1"]

    ws["C3"] = "CCGL9065"
    ws["C4"] = 4252
    ws["C6"] = "Hongshan Guo"
    ws["C8"] = "2A"
    ws["C9"] = "Round"

    max_col = ws.max_column
    template_style_row = 12
    for row in range(12, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None

    output_row = 12
    count = 0
    for source_row in range(2, grades.max_row + 1):
        student_id = sid(grades.cell(source_row, gh["Student #"]).value)
        last = str(grades.cell(source_row, gh["Last Name"]).value or "").strip()
        first = str(grades.cell(source_row, gh["First Name"]).value or "").strip()
        name = f"{first} {last}".strip()

        participation = number(grades.cell(source_row, gh["Tutorial"]).value) + number(
            grades.cell(source_row, gh["Participation"]).value
        )
        assignment = number(grades.cell(source_row, gh["Weekly"]).value)
        portfolio = (
            number(grades.cell(source_row, gh["Essay"]).value)
            + number(grades.cell(source_row, gh["Collage"]).value)
            + number(grades.cell(source_row, gh["Video"]).value)
        )

        coursework_mark = round((participation + assignment) / 40 * 100, 2)
        exam_mark = round(portfolio / 60 * 100, 2)
        roll = rollup_by_id.get(student_id, {})
        final_mark = roll.get("total", round(participation + assignment + portfolio, 2))

        copy_row_style(ws, template_style_row, output_row, max_col)
        ws.cell(output_row, 1).value = count + 1
        ws.cell(output_row, 2).value = int(student_id) if student_id.isdigit() else student_id
        ws.cell(output_row, 3).value = name
        ws.cell(output_row, 4).value = roll.get("letter")
        ws.cell(output_row, 5).value = coursework_mark
        ws.cell(output_row, 6).value = exam_mark
        ws.cell(output_row, 7).value = final_mark
        ws.cell(output_row, 8).value = None
        count += 1
        output_row += 1

    wb.save(OUTPUT)
    print(f"wrote={OUTPUT}")
    print(f"rows={count}")


if __name__ == "__main__":
    main()
