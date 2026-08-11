import collections
import statistics
import time
from pathlib import Path
import shutil

import openpyxl
from openpyxl import load_workbook


WORKBOOK = Path("data/2026/grading_2026_roster_base.xlsx")
WEEKS = ["W2", "W3", "W4", "W5", "W6", "W7", "W8", "W9", "W10", "W11"]


def headers(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}


def ensure_col(ws, name):
    current = headers(ws)
    if name in current:
        return current[name]
    col = ws.max_column + 1
    ws.cell(1, col).value = name
    return col


def is_blank(value):
    return value is None or (
        isinstance(value, str) and value.strip() in ("", "-", "–", "—", "NA", "N/A")
    )


def sid(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def student_name(ws, row, h):
    return (
        str(ws.cell(row, h["First Name"]).value or "").strip()
        + " "
        + str(ws.cell(row, h["Last Name"]).value or "").strip()
    ).strip()


def provisional_band(total):
    if total >= 90:
        return "A range"
    if total >= 80:
        return "B range"
    if total >= 70:
        return "C range"
    if total >= 60:
        return "D range"
    return "F range"


def calibrated_letter(rank):
    if rank <= 3:
        return "A+"
    if rank <= 8:
        return "A"
    if rank <= 17:
        return "A-"
    if rank <= 25:
        return "B+"
    if rank <= 33:
        return "B"
    if rank <= 41:
        return "B-"
    if rank <= 43:
        return "C+"
    if rank <= 45:
        return "C"
    if rank <= 46:
        return "C-"
    return "F"


def recreate_sheet(wb, title):
    if title in wb.sheetnames:
        index = wb.sheetnames.index(title)
        del wb[title]
        return wb.create_sheet(title, index)
    return wb.create_sheet(title)


def set_widths(*worksheets):
    for ws in worksheets:
        for col in range(1, ws.max_column + 1):
            letter = openpyxl.utils.get_column_letter(col)
            header = str(ws.cell(1, col).value or "")
            ws.column_dimensions[letter].width = min(max(12, len(header) + 2), 42)


def main():
    backup = WORKBOOK.with_name(
        f"{WORKBOOK.stem}.before_weekly_median_{time.strftime('%Y%m%d_%H%M%S')}{WORKBOOK.suffix}"
    )
    shutil.copy2(WORKBOOK, backup)

    wb = load_workbook(WORKBOOK)

    weekly = wb["Weekly_Roster"]
    hw = headers(weekly)
    col_median = ensure_col(weekly, "Median_Submitted")
    col_weekly_median = ensure_col(weekly, "Weekly_20_MedianSubmitted")
    col_method = ensure_col(weekly, "Weekly_Method")
    hw = headers(weekly)

    metrics = {}
    blank_total = 0
    week_blanks = collections.Counter()
    blank_distribution = collections.Counter()

    for row in range(2, weekly.max_row + 1):
        values = []
        blanks = []
        for week in WEEKS:
            value = weekly.cell(row, hw[week]).value
            if is_blank(value):
                blanks.append(week)
                blank_total += 1
                week_blanks[week] += 1
                continue
            try:
                values.append(float(value))
            except Exception:
                blanks.append(week)
                blank_total += 1
                week_blanks[week] += 1

        count = len(values)
        raw_total = sum(values)
        median = statistics.median(values) if values else 0.0
        average = raw_total / count if count else 0.0
        weekly_zero_missing = raw_total / 1000 * 20
        weekly_median = median / 100 * 20
        percent_zero_missing = raw_total / 1000 * 100

        blank_distribution[len(blanks)] += 1

        weekly.cell(row, hw["Submitted_Count"]).value = count
        weekly.cell(row, hw["Raw_Total"]).value = round(raw_total, 2)
        weekly.cell(row, hw["Weekly_20_ZeroMissing"]).value = round(weekly_zero_missing, 2)
        weekly.cell(row, hw["Percent_ZeroMissing"]).value = round(percent_zero_missing, 2)
        weekly.cell(row, hw["Percent_SubmittedOnly"]).value = round(average, 2)
        weekly.cell(row, col_median).value = round(median, 2)
        weekly.cell(row, col_weekly_median).value = round(weekly_median, 2)
        weekly.cell(row, col_method).value = (
            "median of non-empty W2-W11 submissions; no submissions = 0"
        )

        key = sid(weekly.cell(row, hw["Student #"]).value)
        metrics[key] = {
            "count": count,
            "raw_total": raw_total,
            "median": median,
            "average": average,
            "weekly_zero_missing": weekly_zero_missing,
            "weekly_median": weekly_median,
            "percent_zero_missing": percent_zero_missing,
            "blank_count": len(blanks),
            "blanks": ",".join(blanks),
            "name": student_name(weekly, row, hw),
        }

    grades = wb["Grades"]
    ensure_col(grades, "Weekly_Median_Submitted")
    ensure_col(grades, "Weekly_Method")
    hg = headers(grades)
    old_weekly_by_key = {}

    for row in range(2, grades.max_row + 1):
        key = sid(grades.cell(row, hg["Student #"]).value)
        metric = metrics.get(key)
        if not metric:
            continue

        old_weekly_by_key[key] = float(grades.cell(row, hg["Weekly"]).value or 0)
        grades.cell(row, hg["Weekly"]).value = round(metric["weekly_median"], 2)
        grades.cell(row, hg["Weekly_Submissions"]).value = metric["count"]
        grades.cell(row, hg["Weekly_Raw_Total"]).value = round(metric["raw_total"], 2)
        grades.cell(row, hg["Weekly_Percent_ZeroMissing"]).value = round(
            metric["percent_zero_missing"], 2
        )
        grades.cell(row, hg["Weekly_Percent_SubmittedOnly"]).value = round(
            metric["average"], 2
        )
        grades.cell(row, hg["Weekly_Median_Submitted"]).value = round(metric["median"], 2)
        grades.cell(row, hg["Weekly_Method"]).value = (
            "median of non-empty W2-W11 submissions; no submissions = 0"
        )

        component_total = sum(
            float(grades.cell(row, hg[component]).value or 0)
            for component in ["Weekly", "Tutorial", "Participation", "Essay", "Collage", "Video"]
        )
        grades.cell(row, hg["Total"]).value = round(component_total, 2)

        name = student_name(grades, row, hg)
        if name in {"Yu-Xuan Chiang", "Hong Kiu Jamie Lee"}:
            previous = grades.cell(row, hg["Manual_Adjustments"]).value
            note = "Prior manual weekly lift superseded by median-submitted weekly policy."
            grades.cell(row, hg["Manual_Adjustments"]).value = (
                f"{previous}; {note}" if previous else note
            )

    final = recreate_sheet(wb, "Final_Rollup")
    final_headers = [
        "Rank",
        "Last Name",
        "First Name",
        "Email",
        "Student #",
        "pXX",
        "Weekly",
        "Tutorial",
        "Participation",
        "Essay",
        "Collage",
        "Video",
        "Computed_Total",
        "Provisional_Band",
        "Calibrated_Letter",
        "Notes",
        "Weekly_Submissions",
        "Weekly_Median_Submitted",
    ]
    for col, header in enumerate(final_headers, 1):
        final.cell(1, col).value = header

    hg = headers(grades)
    ranked = []
    for row in range(2, grades.max_row + 1):
        total = float(grades.cell(row, hg["Total"]).value or 0)
        ranked.append((total, row))
    ranked.sort(
        key=lambda item: (
            -item[0],
            str(grades.cell(item[1], hg["Last Name"]).value or ""),
            str(grades.cell(item[1], hg["First Name"]).value or ""),
        )
    )

    for rank, (total, row) in enumerate(ranked, 1):
        values = [rank]
        for header in [
            "Last Name",
            "First Name",
            "Email",
            "Student #",
            "pXX",
            "Weekly",
            "Tutorial",
            "Participation",
            "Essay",
            "Collage",
            "Video",
        ]:
            values.append(grades.cell(row, hg[header]).value)
        values.extend(
            [
                round(total, 2),
                provisional_band(total),
                calibrated_letter(rank),
                grades.cell(row, hg["Notes"]).value,
                grades.cell(row, hg["Weekly_Submissions"]).value,
                grades.cell(row, hg["Weekly_Median_Submitted"]).value,
            ]
        )
        for col, value in enumerate(values, 1):
            final.cell(rank + 1, col).value = value

    distribution = recreate_sheet(wb, "Distribution_Calibration")
    for col, header in enumerate(["Item", "Value", "Notes"], 1):
        distribution.cell(1, col).value = header
    fh = headers(final)
    letter_counts = collections.Counter(
        final.cell(row, fh["Calibrated_Letter"]).value for row in range(2, final.max_row + 1)
    )
    distribution_rows = [
        (
            "Weekly policy",
            "Median of non-empty W2-W11 submissions, scaled to 20",
            "No weekly submissions remains 0; blank entries no longer count as zero.",
        ),
        ("A+", "Ranks 1-3", f"Count {letter_counts['A+']}"),
        ("A", "Ranks 4-8", f"Count {letter_counts['A']}"),
        ("A-", "Ranks 9-17", f"Count {letter_counts['A-']}"),
        ("B+", "Ranks 18-25", f"Count {letter_counts['B+']}"),
        ("B", "Ranks 26-33", f"Count {letter_counts['B']}"),
        ("B-", "Ranks 34-41", f"Count {letter_counts['B-']}"),
        ("C+", "Ranks 42-43", f"Count {letter_counts['C+']}"),
        ("C", "Ranks 44-45", f"Count {letter_counts['C']}"),
        ("C-", "Rank 46", f"Count {letter_counts['C-']}"),
        ("F", "Ranks 47-48", f"Count {letter_counts['F']}"),
    ]
    for row, values in enumerate(distribution_rows, 2):
        for col, value in enumerate(values, 1):
            distribution.cell(row, col).value = value

    if "Manual_Adjustments" in wb.sheetnames:
        manual = wb["Manual_Adjustments"]
    else:
        manual = wb.create_sheet("Manual_Adjustments")
    if manual.max_row == 1 and all(
        manual.cell(1, col).value is None for col in range(1, manual.max_column + 1)
    ):
        for col, header in enumerate(
            ["Date", "Student", "Component", "Old Value", "New Value", "Reason"], 1
        ):
            manual.cell(1, col).value = header
    for header in ["Date", "Student", "Component", "Old Value", "New Value", "Reason"]:
        ensure_col(manual, header)
    mh = headers(manual)
    row = manual.max_row + 1
    manual.cell(row, mh["Date"]).value = "2026-05-17"
    manual.cell(row, mh["Student"]).value = "All students"
    manual.cell(row, mh["Component"]).value = "Weekly"
    manual.cell(row, mh["Old Value"]).value = (
        "Sum of W2-W11 with blanks as zero, plus two manual weekly lifts"
    )
    manual.cell(row, mh["New Value"]).value = (
        "Median of non-empty W2-W11 submissions scaled to 20; no submissions = 0"
    )
    manual.cell(row, mh["Reason"]).value = (
        "Blank weekly entries appear to over-penalize weekly portfolio quality; "
        "submission count retained for audit."
    )

    qa = recreate_sheet(wb, "QA_Pass")
    for col, header in enumerate(["Check", "Finding", "Detail"], 1):
        qa.cell(1, col).value = header
    impact = []
    for key, metric in metrics.items():
        old = old_weekly_by_key.get(key, metric["weekly_zero_missing"])
        delta = metric["weekly_median"] - old
        if delta >= 5:
            impact.append((delta, metric["name"], metric["count"]))
    impact.sort(reverse=True)
    impact_text = ", ".join(
        f"{name} +{delta:.2f} ({count} submissions)"
        for delta, name, count in impact[:8]
    )
    qa_rows = [
        (
            "Weekly blanks",
            f"{blank_total} blank weekly cells out of {48 * 10}",
            f"W10 is the largest anomaly with {week_blanks['W10']} blanks; blanks are ignored under the new median-submitted policy.",
        ),
        (
            "Weekly policy",
            "Weekly component now uses submitted-work median",
            "Computed as median(non-empty W2-W11 raw scores) / 100 * 20. Students with no submissions receive 0.",
        ),
        (
            "Sparse records",
            "Submission count retained for audit",
            "The policy measures quality of submitted weekly work, not completion frequency; review low-count cases separately if desired.",
        ),
        (
            "Largest weekly lifts",
            impact_text,
            "These are expected consequences of no longer treating blanks as zero.",
        ),
        (
            "Prior manual lifts",
            "Superseded",
            "Yu-Xuan Chiang and Hong Kiu Jamie Lee no longer need one-off weekly treatment under the median-submitted policy.",
        ),
    ]
    for row, values in enumerate(qa_rows, 2):
        for col, value in enumerate(values, 1):
            qa.cell(row, col).value = value

    outliers = recreate_sheet(wb, "QA_Component_Outliers")
    outlier_headers = [
        "Student",
        "Weekly",
        "Weekly_Submissions",
        "Tutorial",
        "Participation",
        "Essay",
        "Collage",
        "Video",
        "Computed_Total",
        "Calibrated_Letter",
        "Flag",
    ]
    for col, header in enumerate(outlier_headers, 1):
        outliers.cell(1, col).value = header
    fh = headers(final)
    outlier_rows = []
    for row in range(2, final.max_row + 1):
        submissions = final.cell(row, fh["Weekly_Submissions"]).value or 0
        weekly_score = final.cell(row, fh["Weekly"]).value or 0
        tutorial = final.cell(row, fh["Tutorial"]).value or 0
        participation = final.cell(row, fh["Participation"]).value or 0
        essay = final.cell(row, fh["Essay"]).value or 0
        collage = final.cell(row, fh["Collage"]).value or 0
        video = final.cell(row, fh["Video"]).value or 0
        total = final.cell(row, fh["Computed_Total"]).value or 0
        letter = final.cell(row, fh["Calibrated_Letter"]).value
        name = student_name(final, row, fh)
        flags = []
        if submissions <= 3 and weekly_score >= 16:
            flags.append("high weekly median from sparse submissions")
        if letter in {"C+", "C", "C-"} and essay + collage + video >= 51:
            flags.append("C range despite near/above-median final portfolio")
        if tutorial == 0 and participation == 0:
            flags.append("tutorial and participation both zero")
        if flags:
            outlier_rows.append(
                (
                    name,
                    weekly_score,
                    submissions,
                    tutorial,
                    participation,
                    essay,
                    collage,
                    video,
                    total,
                    letter,
                    "; ".join(flags),
                )
            )
    for row, values in enumerate(outlier_rows, 2):
        for col, value in enumerate(values, 1):
            outliers.cell(row, col).value = value

    set_widths(weekly, grades, final, distribution, manual, qa, outliers)
    wb.save(WORKBOOK)

    print(f"updated={WORKBOOK}")
    print(f"backup={backup}")
    print(f"weekly_blanks={blank_total}")
    print(f"w10_blanks={week_blanks['W10']}")
    print(f"distribution={dict(sorted(letter_counts.items()))}")


if __name__ == "__main__":
    main()
